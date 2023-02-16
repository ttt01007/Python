# 1.实现自己的Excel数据驱动类
# 2.将写入断言结果做成新的模式
# 3.将描述的字段内容作为日志输出的部分，在运行时随执行行为进行输出
# 4.将已完成的内容，进行封装，在结构上实现优化
#
# 5.非常推荐实现。配置再指定路径下直接读取
# 6.N个文件，包含各种格式，只读取测试用例的格式文件
import logging

import openpyxl
import logging.config
from openpyxl import styles
from openpyxl.styles import PatternFill, Font

from python_selenium.class09_ui_frame.excel_driver import excel_conf
from python_selenium.class09_ui_frame.my_conf import log_conf
from python_selenium.class09_ui_frame.ui_keys.demo01_web_keys import WebKeys


def excel_runner(path, log):
    # logging.config.fileConfig('../my_conf/log.ini', encoding='utf-8')
    # log = log_conf.get_log('../my_conf/log.ini')
    excel = openpyxl.load_workbook(path)
    try:
        sheets = excel.sheetnames
        for sheet in sheets:
            # print(sheet)
            sheet_temp = excel[sheet]
            logging.info('----------{}----------'.format(sheet_temp))
            for values in sheet_temp.values:
                if type(values[0]) is int:
                    # print(values)
                    log.info('正在执行：{}'.format(values[5]))
                    data = {'name': values[2], 'value': values[3], 'text': values[4]}
                    # data = {}
                    # data['name'] = values[2]
                    # data['value'] = values[3]
                    # data['text'] = values[4]

                    # print(data)
                    # print(list(data))
                    # print(data.keys())
                    # print(list(data.keys()))

                    # list(data)与list(data.keys())效果一样
                    for key in list(data.keys()):
                        if data[key] is None:
                            del data[key]
                    # print(data)
                    if values[1] == 'open_browser':
                        wk = WebKeys(values[4], log)
                        sheet_temp.cell(row=values[0] + 2, column=7).value = 'Pass'
                    elif 'assert' in values[1]:
                        status = getattr(wk, values[1])(**data)
                        if status:
                            # sheet_temp.cell(row=values[0] + 2, column=7).value = 'Pass'
                            excel_conf.pass_(sheet_temp.cell, values[0] + 2, 7)
                        else:
                            # sheet_temp.cell(row=values[0] + 2, column=7).value = 'Failed'
                            # sheet_temp.cell(row=values[0] + 2, column=7).fill = PatternFill('solid', fgColor='AACF91')
                            # sheet_temp.cell(row=values[0] + 2, column=7).font = Font(bold=True)
                            excel_conf.falid(sheet_temp.cell, values[0] + 2, 7)
                        excel.save(path)
                    else:
                        getattr(wk, values[1])(**data)
                        sheet_temp.cell(row=values[0] + 2, column=7).value = 'Pass'

        excel.save(path)
    except Exception as e:
        log.error(e)
    finally:
        excel.close()


if __name__ == '__main__':
    excel_runner()
