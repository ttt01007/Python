# 1.实现自己的Excel数据驱动类
# 2.将写入断言结果做成新的模式
# 3.将描述的字段内容作为日志输出的部分，在运行时随执行行为进行输出
# 4.将已完成的内容，进行封装，在结构上实现优化
#
# 5.非常推荐实现。配置再指定路径下直接读取
# 6.N个文件，包含各种格式，只读取测试用例的格式文件
import openpyxl

from python_test.excel.excel_driver.web_keys import WebKeys

excel = openpyxl.load_workbook('../excel_data/data.xlsx')
sheets = excel.sheetnames
for sheet in sheets:
    # print(sheet)
    sheet_temp = excel[sheet]
    for values in sheet_temp.values:
        if type(values[0]) is int:
            # print(values)
            data = {'name': values[2], 'value': values[3], 'text': values[4]}
            # data = {}
            # data['name'] = values[2]
            # data['value'] = values[3]
            # data['text'] = values[4]

            # print(data)
            # print(list(data))
            # print(data.keys())
            # print(list(data.keys()))

            # list(data)与list(data.keys())效果一样
            for key in list(data.keys()):
                if data[key] is None:
                    del data[key]
            print(data)
            if values[1] == 'open_browser':
                wk = WebKeys(values[4])
                sheet_temp.cell(row=values[0] + 2, column=7).value = 'Pass'
            elif 'assert' in values[1]:
                status = getattr(wk, values[1])(**data)
                if status:
                    sheet_temp.cell(row=values[0] + 2, column=7).value = 'Pass'
                else:
                    status: sheet_temp.cell(row=values[0] + 2, column=7).value = 'Failed'
            else:
                getattr(wk, values[1])(**data)
                sheet_temp.cell(row=values[0] + 2, column=7).value = 'Pass'

excel.save('../excel_data/data.xlsx')

# *args 不定长不定值
# **kkwargs 不定长，定值