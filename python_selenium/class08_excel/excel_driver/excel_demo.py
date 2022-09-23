import openpyxl

# 1.读取excel文件
excel = openpyxl.load_workbook('../excel_data/data.xlsx')
# 2.读取sheet页
sheet = excel['Sheet1']
# 3.读取指定的单元格
cell=sheet['A1']
print(cell.value)
# 读取整个 sheet页内容
values=sheet.values
for i in values:
    print(i)

allsheet=excel.sheetnames
for i in allsheet:
    for j in excel[i].values:
        print(j)
        for k in j:
            print(k)

# 写excel
sheet.cell(row=5,column=1).value='qiong'
excel.save('../excel_data/data.xlsx')
